"""Basic openpyxl exporter tests."""
import pytest
from openpyxl import Workbook


def test_create_workbook():
    wb = Workbook()
    assert wb is not None
    ws = wb.active
    assert ws is not None


def test_write_read_cell():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = 42
    assert ws["A1"].value == 42


def test_cell_row_column():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=2, column=3, value="hello")
    assert ws.cell(row=2, column=3).value == "hello"


def test_append_row():
    wb = Workbook()
    ws = wb.active
    ws.append([1, 2, 3])
    assert ws["A1"].value == 1
    assert ws["B1"].value == 2
    assert ws["C1"].value == 3
